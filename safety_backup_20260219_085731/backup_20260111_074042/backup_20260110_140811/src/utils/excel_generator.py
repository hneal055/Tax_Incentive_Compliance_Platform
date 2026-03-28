"""
Excel Export Generator
Professional Excel spreadsheet generation for tax incentive data
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Dict, Any
import io


class ExcelExportGenerator:
    """Generate professional Excel exports"""
    
    def __init__(self):
        # Define color scheme
        self.header_fill = PatternFill(start_color="2c5aa0", end_color="2c5aa0", fill_type="solid")
        self.highlight_fill = PatternFill(start_color="e8f4f8", end_color="e8f4f8", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.title_font = Font(bold=True, size=14)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def _style_header_row(self, ws, row_num):
        """Apply styling to header row"""
        for cell in ws[row_num]:
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border
    
    def _auto_adjust_column_width(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_comparison_workbook(
        self,
        production_title: str,
        budget: float,
        comparisons: List[Dict[str, Any]]
    ) -> bytes:
        """Generate comparison Excel workbook"""
        wb = Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Title
        ws_summary['A1'] = "Tax Incentive Comparison Analysis"
        ws_summary['A1'].font = self.title_font
        ws_summary.merge_cells('A1:E1')
        
        # Production info
        ws_summary['A3'] = "Production:"
        ws_summary['B3'] = production_title
        ws_summary['A4'] = "Total Budget:"
        ws_summary['B4'] = budget
        ws_summary['B4'].number_format = '$#,##0'
        ws_summary['A5'] = "Analysis Date:"
        ws_summary['B5'] = datetime.now().strftime('%B %d, %Y')
        
        # Best option
        if comparisons:
            best = comparisons[0]
            ws_summary['A7'] = "RECOMMENDED LOCATION"
            ws_summary['A7'].font = Font(bold=True, size=12, color="2c5aa0")
            ws_summary['A8'] = "Jurisdiction:"
            ws_summary['B8'] = best['jurisdiction']
            ws_summary['A9'] = "Estimated Credit:"
            ws_summary['B9'] = best['estimatedCredit']
            ws_summary['B9'].number_format = '$#,##0'
            ws_summary['A10'] = "Rate:"
            ws_summary['B10'] = f"{best['percentage']}%" if best['percentage'] else 'N/A'
        
        # Comparison Sheet
        ws_comp = wb.create_sheet("Jurisdictions")
        
        # Headers
        headers = ['Rank', 'Jurisdiction', 'Program', 'Type', 'Rate', 'Estimated Credit', 'Notes']
        ws_comp.append(headers)
        self._style_header_row(ws_comp, 1)
        
        # Data rows
        for comp in comparisons:
            ws_comp.append([
                comp['rank'],
                comp['jurisdiction'],
                comp['ruleName'],
                comp['incentiveType'],
                f"{comp['percentage']}%" if comp['percentage'] else 'N/A',
                comp['estimatedCredit'],
                'Best Option' if comp['rank'] == 1 else ''
            ])
        
        # Format credit column
        for row in range(2, len(comparisons) + 2):
            ws_comp[f'F{row}'].number_format = '$#,##0'
        
        # Highlight best option
        for cell in ws_comp[2]:
            cell.fill = self.highlight_fill
        
        self._auto_adjust_column_width(ws_comp)
        
        # Savings Analysis Sheet
        ws_savings = wb.create_sheet("Savings Analysis")
        
        ws_savings['A1'] = "Savings Comparison"
        ws_savings['A1'].font = self.title_font
        
        if len(comparisons) > 1:
            best = comparisons[0]
            worst = comparisons[-1]
            savings = best['estimatedCredit'] - worst['estimatedCredit']
            
            ws_savings['A3'] = "Best Option:"
            ws_savings['B3'] = best['jurisdiction']
            ws_savings['A4'] = "Best Credit:"
            ws_savings['B4'] = best['estimatedCredit']
            ws_savings['B4'].number_format = '$#,##0'
            
            ws_savings['A6'] = "Lowest Option:"
            ws_savings['B6'] = worst['jurisdiction']
            ws_savings['A7'] = "Lowest Credit:"
            ws_savings['B7'] = worst['estimatedCredit']
            ws_savings['B7'].number_format = '$#,##0'
            
            ws_savings['A9'] = "POTENTIAL SAVINGS:"
            ws_savings['A9'].font = Font(bold=True, color="28a745")
            ws_savings['B9'] = savings
            ws_savings['B9'].number_format = '$#,##0'
            ws_savings['B9'].font = Font(bold=True, size=12, color="28a745")
        
        # Convert to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_compliance_workbook(
        self,
        production_title: str,
        jurisdiction: str,
        rule_name: str,
        requirements: List[Dict[str, Any]],
        overall_status: str,
        estimated_credit: float = None
    ) -> bytes:
        """Generate compliance verification Excel workbook"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Compliance Check"
        
        # Title
        ws['A1'] = "Tax Incentive Compliance Verification"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # Production info
        ws['A3'] = "Production:"
        ws['B3'] = production_title
        ws['A4'] = "Jurisdiction:"
        ws['B4'] = jurisdiction
        ws['A5'] = "Program:"
        ws['B5'] = rule_name
        ws['A6'] = "Analysis Date:"
        ws['B6'] = datetime.now().strftime('%B %d, %Y')
        
        # Overall status
        ws['A8'] = "COMPLIANCE STATUS:"
        ws['A8'].font = Font(bold=True, size=12)
        ws['B8'] = overall_status.upper()
        if overall_status == 'compliant':
            ws['B8'].font = Font(bold=True, size=12, color="28a745")
        else:
            ws['B8'].font = Font(bold=True, size=12, color="dc3545")
        
        if estimated_credit:
            ws['A9'] = "Estimated Credit:"
            ws['B9'] = estimated_credit
            ws['B9'].number_format = '$#,##0'
            ws['B9'].font = Font(bold=True, size=11)
        
        # Requirements table
        ws['A11'] = "Requirements Checklist"
        ws['A11'].font = Font(bold=True, size=11)
        
        headers = ['Status', 'Requirement', 'Description', 'Required']
        ws.append([])  # Row 12
        ws.append(headers)  # Row 13
        self._style_header_row(ws, 13)
        
        for req in requirements:
            status_symbol = '✓' if req['status'] == 'met' else ('?' if req['status'] == 'unknown' else '✗')
            ws.append([
                status_symbol,
                req['requirement'].replace('_', ' ').title(),
                req['description'],
                'Yes' if req.get('required', True) else 'No'
            ])
            
            # Color code status
            last_row = ws.max_row
            if req['status'] == 'met':
                ws[f'A{last_row}'].font = Font(color="28a745", bold=True, size=14)
            elif req['status'] == 'not_met':
                ws[f'A{last_row}'].font = Font(color="dc3545", bold=True, size=14)
            else:
                ws[f'A{last_row}'].font = Font(color="ffc107", bold=True, size=14)
        
        self._auto_adjust_column_width(ws)
        
        # Convert to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_scenario_workbook(
        self,
        production_title: str,
        jurisdiction: str,
        base_budget: float,
        scenarios: List[Dict[str, Any]]
    ) -> bytes:
        """Generate scenario analysis Excel workbook"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Scenario Analysis"
        
        # Title
        ws['A1'] = "Production Scenario Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')
        
        # Info
        ws['A3'] = "Production:"
        ws['B3'] = production_title
        ws['A4'] = "Jurisdiction:"
        ws['B4'] = jurisdiction
        ws['A5'] = "Base Budget:"
        ws['B5'] = base_budget
        ws['B5'].number_format = '$#,##0'
        ws['A6'] = "Scenarios Analyzed:"
        ws['B6'] = len(scenarios)
        
        # Best scenario
        if scenarios:
            best = scenarios[0]
            ws['A8'] = "OPTIMAL SCENARIO"
            ws['A8'].font = Font(bold=True, size=12, color="2c5aa0")
            ws['A9'] = "Scenario:"
            ws['B9'] = best['scenarioName']
            ws['A10'] = "Budget:"
            ws['B10'] = best['scenarioParams']['budget']
            ws['B10'].number_format = '$#,##0'
            ws['A11'] = "Estimated Credit:"
            ws['B11'] = best['estimatedCredit']
            ws['B11'].number_format = '$#,##0'
            ws['A12'] = "Effective Rate:"
            ws['B12'] = f"{best['effectiveRate']:.1f}%"
        
        # Scenarios table
        ws['A14'] = "Scenario Comparison"
        ws['A14'].font = Font(bold=True, size=11)
        
        headers = ['Scenario', 'Budget', 'Program', 'Tax Credit', 'Effective Rate', 'Rank']
        ws.append([])  # Row 15
        ws.append(headers)  # Row 16
        self._style_header_row(ws, 16)
        
        for i, scenario in enumerate(scenarios):
            ws.append([
                scenario['scenarioName'],
                scenario['scenarioParams']['budget'],
                scenario['bestRuleName'],
                scenario['estimatedCredit'],
                scenario['effectiveRate'],
                i + 1
            ])
            
            # Format numbers
            last_row = ws.max_row
            ws[f'B{last_row}'].number_format = '$#,##0'
            ws[f'D{last_row}'].number_format = '$#,##0'
            ws[f'E{last_row}'].number_format = '0.0"%"'
            
            # Highlight best
            if i == 0:
                for col in ['A', 'B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{last_row}'].fill = self.highlight_fill
        
        self._auto_adjust_column_width(ws)
        
        # ROI Analysis Sheet
        ws_roi = wb.create_sheet("ROI Analysis")
        ws_roi['A1'] = "Return on Investment Analysis"
        ws_roi['A1'].font = self.title_font
        
        if len(scenarios) > 1:
            best = scenarios[0]
            worst = scenarios[-1]
            
            ws_roi['A3'] = "Best Scenario:"
            ws_roi['B3'] = best['scenarioName']
            ws_roi['A4'] = "Credit:"
            ws_roi['B4'] = best['estimatedCredit']
            ws_roi['B4'].number_format = '$#,##0'
            
            ws_roi['A6'] = "Worst Scenario:"
            ws_roi['B6'] = worst['scenarioName']
            ws_roi['A7'] = "Credit:"
            ws_roi['B7'] = worst['estimatedCredit']
            ws_roi['B7'].number_format = '$#,##0'
            
            optimization = best['estimatedCredit'] - worst['estimatedCredit']
            ws_roi['A9'] = "OPTIMIZATION POTENTIAL:"
            ws_roi['A9'].font = Font(bold=True, color="28a745")
            ws_roi['B9'] = optimization
            ws_roi['B9'].number_format = '$#,##0'
            ws_roi['B9'].font = Font(bold=True, size=12, color="28a745")
        
        # Convert to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


# Global Excel generator instance
excel_generator = ExcelExportGenerator()