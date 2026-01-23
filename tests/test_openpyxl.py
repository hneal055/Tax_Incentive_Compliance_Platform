import openpyxl
from openpyxl import Workbook

# Create a new workbook
wb = Workbook()
ws = wb.active

# Add some data
ws['A1'] = 'Tax Incentive'
ws['B1'] = 'Amount'
ws['A2'] = 'R&D Credit'
ws['B2'] = 45000
ws['A3'] = 'Energy Credit'
ws['B3'] = 32500

# Save the file
wb.save('test_excel.xlsx')
print('✅ Excel file created: test_excel.xlsx')
print(f'✅ openpyxl version: {openpyxl.__version__}')
